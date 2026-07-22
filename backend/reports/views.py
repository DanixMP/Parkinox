"""
Views for the reports app.

Requirements: 44.1, 44.2, 44.4, 44.5, 44.6, 44.7,
              45.1, 45.2, 45.3, 45.4, 45.5,
              53.4, 82.1-82.7, 83.1-83.7, 84.1-84.5
"""

import io
import logging
from datetime import date
from decimal import Decimal

from django.db.models import Sum, Avg, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

from parking.models import ParkingSession
from utils.tehran_dates import TEHRAN_TZ, date_range_tehran, parse_date_param
from .permissions import IsOperatorOrAdmin

logger = logging.getLogger(__name__)

# Persian translation maps
PAYMENT_METHOD_FA = {
    'wallet': 'کیف پول',
    'cash': 'نقدی',
    'waived': 'بخشوده',
}

STATUS_FA = {
    'parked': 'در پارک',
    'completed': 'تکمیل',
    'unpaid': 'پرداخت نشده',
    'waived': 'بخشوده',
    'flagged': 'علامت‌گذاری',
}


def _parse_date_param(request):
    """Parse and validate the ?date=YYYY-MM-DD query parameter."""
    return parse_date_param(request, required=True)


def _date_range_utc(requested_date: date):
    """Return (start, end) for the given date in Asia/Tehran timezone."""
    return date_range_tehran(requested_date)


def _build_summary(requested_date: date) -> dict:
    """
    Build the daily summary dict for the given date.
    Requirements: 45.1, 45.2, 45.3, 45.4, 45.5, 84.1-84.5
    """
    start_utc, end_utc = _date_range_utc(requested_date)

    # Base queryset: sessions whose entry_time falls within the date
    day_sessions = ParkingSession.objects.filter(
        entry_time__gte=start_utc,
        entry_time__lt=end_utc,
    )

    total_entries = day_sessions.count()

    total_exits = day_sessions.filter(
        exit_time__isnull=False,
        status__in=('completed', 'unpaid', 'waived'),
    ).count()

    revenue_qs = day_sessions.filter(status__in=('completed', 'waived'))
    total_revenue = revenue_qs.aggregate(s=Sum('fee_charged'))['s'] or Decimal('0')

    revenue_by_wallet = (
        day_sessions.filter(payment_method='wallet')
        .aggregate(s=Sum('fee_charged'))['s'] or Decimal('0')
    )
    revenue_by_cash = (
        day_sessions.filter(payment_method='cash')
        .aggregate(s=Sum('fee_charged'))['s'] or Decimal('0')
    )
    revenue_by_waived = (
        day_sessions.filter(status='waived')
        .aggregate(s=Sum('fee_charged'))['s'] or Decimal('0')
    )

    waived_count = day_sessions.filter(status='waived').count()

    avg_duration = (
        day_sessions.filter(status='completed', duration_minutes__isnull=False)
        .aggregate(a=Avg('duration_minutes'))['a']
    )
    average_duration_minutes = round(float(avg_duration), 1) if avg_duration is not None else None

    currently_parked = ParkingSession.objects.filter(status='parked').count()

    return {
        'date': str(requested_date),
        'total_entries': total_entries,
        'total_exits': total_exits,
        'total_revenue': int(total_revenue),
        'revenue_by_wallet': int(revenue_by_wallet),
        'revenue_by_cash': int(revenue_by_cash),
        'revenue_by_waived': int(revenue_by_waived),
        'waived_count': waived_count,
        'average_duration_minutes': average_duration_minutes,
        'currently_parked': currently_parked,
    }


def _get_sessions_for_date(requested_date: date):
    """Return ParkingSession queryset for the given date (Tehran timezone)."""
    start_utc, end_utc = _date_range_utc(requested_date)
    return ParkingSession.objects.filter(
        entry_time__gte=start_utc,
        entry_time__lt=end_utc,
    ).select_related('plate', 'plate__user').order_by('entry_time')


def _format_dt(dt) -> str:
    """Format a UTC datetime as Asia/Tehran local time string."""
    if dt is None:
        return ''
    local_dt = dt.astimezone(TEHRAN_TZ)
    return local_dt.strftime('%Y-%m-%d %H:%M:%S')


def _format_fee(fee) -> str:
    """Format fee with thousand separators."""
    if fee is None:
        return '0'
    return f'{int(fee):,}'


def _plate_display(session) -> str:
    """Return the plate number string for a session."""
    if session.plate:
        return session.plate.plate_number
    return session.guest_plate_number or ''


def _serialize_session_row(session, index: int) -> dict:
    """JSON row for a parking session on a given day."""
    payment_method = session.payment_method or ''
    return {
        'row': index,
        'id': session.id,
        'plate_number': _plate_display(session),
        'user_name': (
            session.plate.user.full_name
            if session.plate_id and session.plate
            else 'مهمان'
        ),
        'entry_time': _format_dt(session.entry_time),
        'exit_time': _format_dt(session.exit_time),
        'duration_minutes': session.duration_minutes,
        'fee_charged': int(session.fee_charged or 0),
        'payment_method': payment_method,
        'payment_method_fa': PAYMENT_METHOD_FA.get(payment_method, payment_method),
        'status': session.status,
        'status_fa': STATUS_FA.get(session.status, session.status),
        'is_registered': session.plate_id is not None,
    }


# ---------------------------------------------------------------------------
# View: DailySessionsView
# ---------------------------------------------------------------------------

class DailySessionsView(APIView):
    """
    GET /api/reports/daily/sessions/?date=YYYY-MM-DD

    Returns session list + summary for archive UI.
    """

    permission_classes = [IsAuthenticated, IsOperatorOrAdmin]

    def get(self, request):
        requested_date, err = _parse_date_param(request)
        if err:
            return err

        try:
            sessions = list(_get_sessions_for_date(requested_date))
            rows = [
                _serialize_session_row(s, idx)
                for idx, s in enumerate(sessions, start=1)
            ]
            summary = _build_summary(requested_date)
            return Response(
                {
                    'date': str(requested_date),
                    'summary': summary,
                    'count': len(rows),
                    'sessions': rows,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as exc:
            logger.error('Error building daily sessions: %s', exc, exc_info=True)
            return Response(
                {'error': 'Internal server error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ---------------------------------------------------------------------------
# View: DailySummaryView
# ---------------------------------------------------------------------------

class DailySummaryView(APIView):
    """
    GET /api/reports/daily/?date=YYYY-MM-DD

    Returns daily parking statistics for the given date.

    Requirements: 45.1, 45.2, 45.3, 45.4, 45.5, 53.4, 84.1, 84.2, 84.3, 84.4, 84.5
    """

    permission_classes = [IsAuthenticated, IsOperatorOrAdmin]

    def get(self, request):
        requested_date, err = _parse_date_param(request)
        if err:
            return err

        try:
            summary = _build_summary(requested_date)
            return Response(summary, status=status.HTTP_200_OK)
        except Exception as exc:
            logger.error('Error building daily summary: %s', exc, exc_info=True)
            return Response(
                {'error': 'Internal server error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ---------------------------------------------------------------------------
# View: ExcelReportView
# ---------------------------------------------------------------------------

class ExcelReportView(APIView):
    """
    GET /api/reports/export/excel/?date=YYYY-MM-DD

    Generates and returns an Excel (.xlsx) report for the given date.

    Requirements: 44.1, 44.4, 44.7, 82.1, 82.2, 82.3, 82.4, 82.5, 82.6, 82.7
    """

    permission_classes = [IsAuthenticated, IsOperatorOrAdmin]

    def get(self, request):
        requested_date, err = _parse_date_param(request)
        if err:
            return err

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'openpyxl is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            wb = openpyxl.Workbook()

            # ---- Sheet 1: Sessions ----
            ws_sessions = wb.active
            ws_sessions.title = 'Sessions'

            headers = [
                'شماره',       # Row
                'پلاک',        # Plate
                'زمان ورود',   # Entry Time
                'زمان خروج',   # Exit Time
                'مدت - دقیقه', # Duration
                'مبلغ - تومان',# Fee
                'روش پرداخت',  # Payment Method
                'وضعیت',       # Status
            ]

            # Write header row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_sessions.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            # Freeze top row
            ws_sessions.freeze_panes = 'A2'

            sessions = _get_sessions_for_date(requested_date)

            for row_idx, session in enumerate(sessions, start=2):
                ws_sessions.cell(row=row_idx, column=1, value=row_idx - 1)
                ws_sessions.cell(row=row_idx, column=2, value=_plate_display(session))
                ws_sessions.cell(row=row_idx, column=3, value=_format_dt(session.entry_time))
                ws_sessions.cell(row=row_idx, column=4, value=_format_dt(session.exit_time))
                ws_sessions.cell(row=row_idx, column=5, value=session.duration_minutes)
                ws_sessions.cell(row=row_idx, column=6, value=_format_fee(session.fee_charged))
                ws_sessions.cell(
                    row=row_idx, column=7,
                    value=PAYMENT_METHOD_FA.get(session.payment_method or '', '')
                )
                ws_sessions.cell(
                    row=row_idx, column=8,
                    value=STATUS_FA.get(session.status, session.status)
                )

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(headers[col_idx - 1])
                for row in ws_sessions.iter_rows(
                    min_row=1, max_row=ws_sessions.max_row,
                    min_col=col_idx, max_col=col_idx
                ):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws_sessions.column_dimensions[col_letter].width = adjusted_width

            # ---- Sheet 2: Summary ----
            ws_summary = wb.create_sheet(title='Summary')
            summary = _build_summary(requested_date)

            summary_rows = [
                ('تاریخ', summary['date']),
                ('کل جلسات', summary['total_entries']),
                ('کل ورودی‌ها', summary['total_entries']),
                ('کل خروجی‌ها', summary['total_exits']),
                ('درآمد کل (تومان)', summary['total_revenue']),
                ('درآمد کیف پول (تومان)', summary['revenue_by_wallet']),
                ('درآمد نقدی (تومان)', summary['revenue_by_cash']),
                ('تعداد بخشوده', summary['waived_count']),
                ('میانگین مدت (دقیقه)', summary['average_duration_minutes']),
                ('در حال پارک', summary['currently_parked']),
            ]

            ws_summary.cell(row=1, column=1, value='شاخص').font = Font(bold=True)
            ws_summary.cell(row=1, column=2, value='مقدار').font = Font(bold=True)

            for row_idx, (label, value) in enumerate(summary_rows, start=2):
                ws_summary.cell(row=row_idx, column=1, value=label)
                ws_summary.cell(row=row_idx, column=2, value=value)

            # Auto-size summary columns
            for col_idx in (1, 2):
                col_letter = get_column_letter(col_idx)
                max_length = 10
                for row in ws_summary.iter_rows(
                    min_row=1, max_row=ws_summary.max_row,
                    min_col=col_idx, max_col=col_idx
                ):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws_summary.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f'parkinox_report_{requested_date}.xlsx'
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as exc:
            logger.error('Error generating Excel report: %s', exc, exc_info=True)
            return Response(
                {'error': 'Internal server error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ---------------------------------------------------------------------------
# View: PDFReportView
# ---------------------------------------------------------------------------

class PDFReportView(APIView):
    """
    GET /api/reports/export/pdf/?date=YYYY-MM-DD

    Generates and returns a PDF report for the given date.

    Requirements: 44.2, 44.5, 44.6, 44.7, 83.1, 83.2, 83.3, 83.4, 83.5, 83.6, 83.7
    """

    permission_classes = [IsAuthenticated, IsOperatorOrAdmin]

    def get(self, request):
        requested_date, err = _parse_date_param(request)
        if err:
            return err

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer, PageBreak,
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            return Response(
                {'error': 'reportlab is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            # ---- Font setup (graceful degradation) ----
            import os
            from pathlib import Path

            BASE_DIR = Path(__file__).resolve().parent.parent
            font_path = BASE_DIR / 'assets' / 'fonts' / 'Vazirmatn-Regular.ttf'
            use_persian_font = False
            font_name = 'Helvetica'

            if font_path.exists():
                try:
                    pdfmetrics.registerFont(TTFont('Vazirmatn', str(font_path)))
                    font_name = 'Vazirmatn'
                    use_persian_font = True
                except Exception:
                    pass

            # ---- Arabic reshaper / bidi (graceful degradation) ----
            def reshape_text(text: str) -> str:
                if not use_persian_font:
                    return text
                try:
                    import arabic_reshaper
                    from bidi.algorithm import get_display
                    reshaped = arabic_reshaper.reshape(text)
                    return get_display(reshaped)
                except ImportError:
                    return text

            # ---- Page number canvas ----
            buffer = io.BytesIO()

            # We need page numbers, so we use a custom canvas factory
            page_count_holder = [0]

            class _NumberedCanvas:
                """Mixin-style page number support via onLaterPages / onFirstPage."""
                pass

            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=2 * cm,
                leftMargin=2 * cm,
                topMargin=2 * cm,
                bottomMargin=2 * cm,
            )

            styles = getSampleStyleSheet()
            normal_style = ParagraphStyle(
                'NormalCustom',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
            )
            title_style = ParagraphStyle(
                'TitleCustom',
                parent=styles['Title'],
                fontName=font_name,
                fontSize=16,
                spaceAfter=12,
            )
            heading_style = ParagraphStyle(
                'HeadingCustom',
                parent=styles['Heading2'],
                fontName=font_name,
                fontSize=12,
                spaceAfter=6,
            )

            story = []

            # ---- Header section ----
            report_title = reshape_text('Parkinox Daily Report')
            story.append(Paragraph(report_title, title_style))

            now_tehran = timezone.now().astimezone(TEHRAN_TZ)
            story.append(Paragraph(
                f"Date: {requested_date}",
                normal_style,
            ))
            story.append(Paragraph(
                f"Generated at: {now_tehran.strftime('%Y-%m-%d %H:%M:%S')} (Asia/Tehran)",
                normal_style,
            ))
            story.append(Spacer(1, 0.5 * cm))

            # ---- Summary table ----
            summary = _build_summary(requested_date)

            story.append(Paragraph(reshape_text('Summary / خلاصه'), heading_style))

            summary_data = [
                [reshape_text('شاخص'), reshape_text('مقدار')],
                [reshape_text('تاریخ'), str(summary['date'])],
                [reshape_text('کل ورودی‌ها'), str(summary['total_entries'])],
                [reshape_text('کل خروجی‌ها'), str(summary['total_exits'])],
                [reshape_text('درآمد کل (تومان)'), f"{summary['total_revenue']:,}"],
                [reshape_text('درآمد کیف پول (تومان)'), f"{summary['revenue_by_wallet']:,}"],
                [reshape_text('درآمد نقدی (تومان)'), f"{summary['revenue_by_cash']:,}"],
                [reshape_text('تعداد بخشوده'), str(summary['waived_count'])],
                [reshape_text('میانگین مدت (دقیقه)'), str(summary['average_duration_minutes'])],
                [reshape_text('در حال پارک'), str(summary['currently_parked'])],
            ]

            summary_table = Table(summary_data, colWidths=[10 * cm, 6 * cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003087')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.5 * cm))

            # ---- Sessions table ----
            story.append(Paragraph(reshape_text('Sessions / جلسات'), heading_style))

            col_headers = [
                reshape_text('شماره'),
                reshape_text('پلاک'),
                reshape_text('زمان ورود'),
                reshape_text('زمان خروج'),
                reshape_text('مدت'),
                reshape_text('مبلغ'),
                reshape_text('روش پرداخت'),
                reshape_text('وضعیت'),
            ]

            sessions = list(_get_sessions_for_date(requested_date))
            table_data = [col_headers]

            for idx, session in enumerate(sessions, start=1):
                table_data.append([
                    str(idx),
                    _plate_display(session),
                    _format_dt(session.entry_time),
                    _format_dt(session.exit_time),
                    str(session.duration_minutes or ''),
                    _format_fee(session.fee_charged),
                    reshape_text(PAYMENT_METHOD_FA.get(session.payment_method or '', '')),
                    reshape_text(STATUS_FA.get(session.status, session.status)),
                ])

            # Column widths (total ~17cm usable width on A4 with 2cm margins each side)
            col_widths = [1.0 * cm, 3.0 * cm, 3.5 * cm, 3.5 * cm,
                          1.5 * cm, 2.0 * cm, 2.5 * cm, 2.5 * cm]

            sessions_table = Table(table_data, colWidths=col_widths, repeatRows=1)

            # Build alternating row colors
            row_styles = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003087')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            for row_idx in range(1, len(table_data)):
                bg = colors.white if row_idx % 2 == 1 else colors.HexColor('#F5F5F5')
                row_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg))

            sessions_table.setStyle(TableStyle(row_styles))
            story.append(sessions_table)

            # ---- Build PDF with page numbers ----
            total_pages_holder = {'total': 1}

            def _add_page_number(canvas, doc):
                canvas.saveState()
                page_num = canvas.getPageNumber()
                total_pages_holder['total'] = max(total_pages_holder['total'], page_num)
                text = f'Page {page_num}'
                canvas.setFont(font_name, 8)
                canvas.drawRightString(
                    A4[0] - 2 * cm,
                    1.5 * cm,
                    text,
                )
                canvas.restoreState()

            doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)

            buffer.seek(0)
            filename = f'parkinox_report_{requested_date}.pdf'
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as exc:
            logger.error('Error generating PDF report: %s', exc, exc_info=True)
            return Response(
                {'error': 'Internal server error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

"""
Unit tests for the reports app.

Requirements: 45.4, 54.7, 84.4
"""

import time
from datetime import datetime, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

from django.test import TestCase
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from accounts.models import User
from plates.models import Plate
from parking.models import ParkingSession

TEHRAN_TZ = ZoneInfo('Asia/Tehran')


def _tehran_dt(year, month, day, hour=0, minute=0, second=0):
    """Return a timezone-aware datetime in Asia/Tehran."""
    return datetime(year, month, day, hour, minute, second, tzinfo=TEHRAN_TZ)


class ReportsTestCase(TestCase):
    """
    Base test case that sets up common fixtures for all report tests.
    """

    def setUp(self):
        self.client = APIClient()

        # Operator user (has access to reports)
        self.operator = User.objects.create_user(
            phone='+989120000001',
            password='testpass123',
            full_name='Test Operator',
            role='operator',
        )

        # Student user (no access to reports)
        self.student = User.objects.create_user(
            phone='+989120000002',
            password='testpass123',
            full_name='Test Student',
            student_id='S001',
            role='student',
        )

        # Plate for student
        self.plate = Plate.objects.create(
            user=self.student,
            plate_number='۱۲ب۳۴۵-۶۷',
            is_primary=True,
            is_verified=True,
            is_active=True,
        )

        # Test date: 2024-03-01 (a fixed date for reproducible tests)
        self.test_date = '2024-03-01'
        self.test_date_obj = datetime(2024, 3, 1).date()

        # Entry time at 09:00 Tehran on 2024-03-01
        self.entry_time = _tehran_dt(2024, 3, 1, 9, 0, 0)
        # Exit time at 11:30 Tehran on 2024-03-01 (150 minutes)
        self.exit_time = _tehran_dt(2024, 3, 1, 11, 30, 0)

        # Authenticate as operator by default
        self.client.force_authenticate(user=self.operator)

        self.daily_url = '/api/reports/daily/'
        self.excel_url = '/api/reports/export/excel/'
        self.pdf_url = '/api/reports/export/pdf/'

    def _create_session(self, **kwargs):
        """
        Helper to create a ParkingSession bypassing full_clean validation
        for fields we don't need to set in tests.
        """
        defaults = {
            'plate': self.plate,
            'entry_time': self.entry_time,
            'fee_charged': Decimal('10000'),
            'status': 'completed',
            'payment_method': 'wallet',
            'duration_minutes': 150,
            'exit_time': self.exit_time,
        }
        defaults.update(kwargs)
        # Use save(force_insert=True) to bypass the custom save() that calls full_clean
        session = ParkingSession(**defaults)
        # Call full_clean manually to validate, then save
        session.full_clean()
        session.save()
        return session


# ---------------------------------------------------------------------------
# Test 1: Daily summary returns correct counts
# ---------------------------------------------------------------------------

class DailySummaryCountsTest(ReportsTestCase):
    """
    Test that daily summary returns correct entry/exit counts.
    Requirements: 45.1, 45.2, 84.1, 84.2
    """

    def test_daily_summary_returns_correct_counts(self):
        """Create sessions for a date and verify entry/exit counts."""
        # Create 3 completed sessions on test date
        for i in range(3):
            self._create_session(
                entry_time=_tehran_dt(2024, 3, 1, 9 + i, 0, 0),
                exit_time=_tehran_dt(2024, 3, 1, 10 + i, 0, 0),
                duration_minutes=60,
                fee_charged=Decimal('5000'),
                status='completed',
                payment_method='wallet',
            )

        # Create 1 still-parked session (no exit)
        self._create_session(
            entry_time=_tehran_dt(2024, 3, 1, 14, 0, 0),
            exit_time=None,
            duration_minutes=None,
            fee_charged=Decimal('0'),
            status='parked',
            payment_method=None,
        )

        response = self.client.get(self.daily_url, {'date': self.test_date})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data
        self.assertEqual(data['total_entries'], 4)
        self.assertEqual(data['total_exits'], 3)
        self.assertEqual(data['currently_parked'], 1)


# ---------------------------------------------------------------------------
# Test 2: Revenue calculation
# ---------------------------------------------------------------------------

class DailySummaryRevenueTest(ReportsTestCase):
    """
    Test that revenue sums correctly by payment method.
    Requirements: 45.3, 84.3
    """

    def test_daily_summary_revenue_calculation(self):
        """Verify revenue sums correctly by payment method."""
        # Wallet payment: 10,000
        self._create_session(
            fee_charged=Decimal('10000'),
            status='completed',
            payment_method='wallet',
        )
        # Cash payment: 20,000
        self._create_session(
            entry_time=_tehran_dt(2024, 3, 1, 10, 0, 0),
            exit_time=_tehran_dt(2024, 3, 1, 12, 0, 0),
            fee_charged=Decimal('20000'),
            status='completed',
            payment_method='cash',
        )
        # Waived session: fee = 0
        self._create_session(
            entry_time=_tehran_dt(2024, 3, 1, 13, 0, 0),
            exit_time=_tehran_dt(2024, 3, 1, 14, 0, 0),
            fee_charged=Decimal('0'),
            status='waived',
            payment_method='waived',
            duration_minutes=60,
        )

        response = self.client.get(self.daily_url, {'date': self.test_date})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data
        self.assertEqual(data['total_revenue'], 30000)
        self.assertEqual(data['revenue_by_wallet'], 10000)
        self.assertEqual(data['revenue_by_cash'], 20000)
        self.assertEqual(data['revenue_by_waived'], 0)
        self.assertEqual(data['waived_count'], 1)


# ---------------------------------------------------------------------------
# Test 3: Invalid date format
# ---------------------------------------------------------------------------

class DailySummaryInvalidDateTest(ReportsTestCase):
    """
    Test that invalid date format returns 400.
    Requirements: 45.5, 84.5
    """

    def test_daily_summary_invalid_date_format(self):
        """Verify 400 returned for bad date."""
        response = self.client.get(self.daily_url, {'date': 'not-a-date'})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_daily_summary_missing_date_param(self):
        """Verify 400 returned when date param is missing."""
        response = self.client.get(self.daily_url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_daily_summary_wrong_date_format(self):
        """Verify 400 returned for DD/MM/YYYY format."""
        response = self.client.get(self.daily_url, {'date': '01/03/2024'})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


# ---------------------------------------------------------------------------
# Test 4: Authentication required
# ---------------------------------------------------------------------------

class DailySummaryAuthTest(ReportsTestCase):
    """
    Test that endpoints require authentication.
    Requirements: 76.1, 76.2
    """

    def test_daily_summary_requires_authentication(self):
        """Verify 401 without token."""
        self.client.force_authenticate(user=None)
        response = self.client.get(self.daily_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_excel_export_requires_authentication(self):
        """Verify 401 without token for Excel export."""
        self.client.force_authenticate(user=None)
        response = self.client.get(self.excel_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_pdf_export_requires_authentication(self):
        """Verify 401 without token for PDF export."""
        self.client.force_authenticate(user=None)
        response = self.client.get(self.pdf_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


# ---------------------------------------------------------------------------
# Test 5: Role-based access control
# ---------------------------------------------------------------------------

class DailySummaryRoleTest(ReportsTestCase):
    """
    Test that student role cannot access reports.
    Requirements: 2.5, 77.3
    """

    def test_daily_summary_requires_operator_role(self):
        """Verify 403 for student role."""
        self.client.force_authenticate(user=self.student)
        response = self.client.get(self.daily_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_excel_export_requires_operator_role(self):
        """Verify 403 for student role on Excel export."""
        self.client.force_authenticate(user=self.student)
        response = self.client.get(self.excel_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_pdf_export_requires_operator_role(self):
        """Verify 403 for student role on PDF export."""
        self.client.force_authenticate(user=self.student)
        response = self.client.get(self.pdf_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


# ---------------------------------------------------------------------------
# Test 6: Excel export returns xlsx
# ---------------------------------------------------------------------------

class ExcelExportTest(ReportsTestCase):
    """
    Test Excel export content-type and filename.
    Requirements: 82.1, 82.2, 82.3
    """

    def test_excel_export_returns_xlsx(self):
        """Verify content-type and filename header."""
        response = self.client.get(self.excel_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            f'parkinox_report_{self.test_date}.xlsx',
            response['Content-Disposition'],
        )

    def test_excel_export_has_correct_sheets(self):
        """Load workbook and verify sheet names."""
        import io
        import openpyxl

        response = self.client.get(self.excel_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn('Sessions', wb.sheetnames)
        self.assertIn('Summary', wb.sheetnames)

    def test_excel_export_sessions_sheet_has_headers(self):
        """Verify Sessions sheet has the expected header row."""
        import io
        import openpyxl

        response = self.client.get(self.excel_url, {'date': self.test_date})
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['Sessions']

        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        self.assertIn('پلاک', headers)
        self.assertIn('زمان ورود', headers)
        self.assertIn('وضعیت', headers)


# ---------------------------------------------------------------------------
# Test 7: PDF export returns pdf
# ---------------------------------------------------------------------------

class PDFExportTest(ReportsTestCase):
    """
    Test PDF export content-type and filename.
    Requirements: 83.1, 83.2, 83.3
    """

    def test_pdf_export_returns_pdf(self):
        """Verify content-type and filename header."""
        response = self.client.get(self.pdf_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn(
            f'parkinox_report_{self.test_date}.pdf',
            response['Content-Disposition'],
        )

    def test_pdf_export_content_is_valid_pdf(self):
        """Verify the response body starts with the PDF magic bytes."""
        response = self.client.get(self.pdf_url, {'date': self.test_date})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(b'%PDF'))


# ---------------------------------------------------------------------------
# Test 8: Response time
# ---------------------------------------------------------------------------

class DailySummaryResponseTimeTest(ReportsTestCase):
    """
    Test that daily summary responds within 2 seconds.
    Requirements: 53.4, 84.4
    """

    def test_daily_summary_response_time(self):
        """Verify response under 2 seconds."""
        start = time.time()
        response = self.client.get(self.daily_url, {'date': self.test_date})
        elapsed = time.time() - start

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertLess(elapsed, 2.0, f'Response took {elapsed:.3f}s, expected < 2s')


# ---------------------------------------------------------------------------
# Test 9: Empty date returns zeros
# ---------------------------------------------------------------------------

class DailySummaryEmptyDateTest(ReportsTestCase):
    """
    Test that a date with no sessions returns zeros.
    Requirements: 84.1, 84.2, 84.3
    """

    def test_daily_summary_empty_date(self):
        """Verify zeros returned for date with no sessions."""
        response = self.client.get(self.daily_url, {'date': '2000-01-01'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data
        self.assertEqual(data['total_entries'], 0)
        self.assertEqual(data['total_exits'], 0)
        self.assertEqual(data['total_revenue'], 0)
        self.assertEqual(data['revenue_by_wallet'], 0)
        self.assertEqual(data['revenue_by_cash'], 0)
        self.assertEqual(data['waived_count'], 0)
        self.assertIsNone(data['average_duration_minutes'])

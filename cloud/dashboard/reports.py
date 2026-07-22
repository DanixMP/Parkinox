"""Excel / PDF / CSV report builders for cloud replica data."""

from __future__ import annotations

import csv
import io
from datetime import datetime

from django.db.models import Avg, Sum
from django.utils import timezone

from replica.models import ReplicaGateEvent, ReplicaParkingSession

from .services import filter_sessions, overview_stats


def build_csv_sessions(site, params) -> bytes:
    qs = filter_sessions(site, params)[:5000]
    buf = io.StringIO()
    buf.write('\ufeff')  # BOM for Excel
    w = csv.writer(buf)
    w.writerow(
        [
            'plate',
            'status',
            'entry_time',
            'exit_time',
            'duration_minutes',
            'fee',
            'payment_method',
            'entry_camera',
            'exit_camera',
            'user_name',
            'session_uuid',
        ]
    )
    for s in qs:
        w.writerow(
            [
                s.plate_number,
                s.status,
                s.entry_time.isoformat() if s.entry_time else '',
                s.exit_time.isoformat() if s.exit_time else '',
                s.duration_minutes or '',
                s.fee_charged,
                s.payment_method,
                s.entry_camera_id,
                s.exit_camera_id,
                s.user_name,
                str(s.session_uuid),
            ]
        )
    return buf.getvalue().encode('utf-8')


def build_excel(site, report_type: str, params) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    stats = overview_stats(site)
    ws.append(['Parkinox Cloud Report'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Site', site.site_id if site else ''])
    ws.append(['Generated', timezone.localtime().isoformat()])
    ws.append([])
    for k, v in stats.items():
        ws.append([k, str(v)])

    ws2 = wb.create_sheet('Sessions')
    qs = filter_sessions(site, params)[:5000]
    if report_type == 'parked':
        qs = ReplicaParkingSession.objects.filter(site=site, status='parked').order_by('entry_time')
    headers = [
        'plate',
        'status',
        'entry_time',
        'exit_time',
        'duration',
        'fee',
        'payment',
        'entry_gate',
        'exit_gate',
        'user',
        'session_uuid',
    ]
    ws2.append(headers)
    for s in qs:
        ws2.append(
            [
                s.plate_number,
                s.status,
                timezone.localtime(s.entry_time).strftime('%Y-%m-%d %H:%M') if s.entry_time else '',
                timezone.localtime(s.exit_time).strftime('%Y-%m-%d %H:%M') if s.exit_time else '',
                s.duration_minutes,
                int(s.fee_charged or 0),
                s.payment_method,
                s.entry_camera_id,
                s.exit_camera_id,
                s.user_name,
                str(s.session_uuid),
            ]
        )

    ws3 = wb.create_sheet('GateEvents')
    events = ReplicaGateEvent.objects.filter(site=site).order_by('-occurred_at')[:5000]
    ws3.append(
        [
            'event_uuid',
            'direction',
            'plate',
            'camera',
            'confidence',
            'edited',
            'auto',
            'occurred_at',
        ]
    )
    for e in events:
        ws3.append(
            [
                str(e.event_uuid),
                e.direction,
                e.plate_confirmed,
                e.camera_id,
                e.confidence,
                e.was_edited,
                e.was_auto_approved,
                e.occurred_at.isoformat() if e.occurred_at else '',
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_pdf(site, report_type: str, params) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    stats = overview_stats(site)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph('Parkinox Management Summary', styles['Title']))
    story.append(Paragraph(f"Site: {site.site_id if site else '-'}", styles['Normal']))
    story.append(Paragraph(f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 16))

    data = [
        ['Metric', 'Value'],
        ['Currently parked', str(stats.get('parked'))],
        ['Capacity', str(stats.get('capacity'))],
        ['Available', str(stats.get('available'))],
        ['Today entries', str(stats.get('today_entries'))],
        ['Today exits', str(stats.get('today_exits'))],
        ['Today revenue (Toman)', f"{stats.get('today_revenue'):,}"],
        ['Avg duration (min)', str(stats.get('avg_duration'))],
    ]
    table = Table(data, colWidths=[220, 220])
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.95, 0.96, 0.98)]),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 20))

    qs = list(filter_sessions(site, params)[:40])
    rows = [['Plate', 'Status', 'Entry', 'Exit', 'Fee']]
    for s in qs:
        rows.append(
            [
                s.plate_number,
                s.status,
                timezone.localtime(s.entry_time).strftime('%m-%d %H:%M') if s.entry_time else '',
                timezone.localtime(s.exit_time).strftime('%m-%d %H:%M') if s.exit_time else '',
                str(int(s.fee_charged or 0)),
            ]
        )
    t2 = Table(rows, colWidths=[90, 70, 90, 90, 70])
    t2.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ]
        )
    )
    story.append(Paragraph('Recent sessions (compact)', styles['Heading2']))
    story.append(t2)
    story.append(Spacer(1, 12))
    story.append(
        Paragraph(
            'Full-resolution images are not embedded. Download image ZIP from session detail.',
            styles['Italic'],
        )
    )
    doc.build(story)
    return buf.getvalue()
